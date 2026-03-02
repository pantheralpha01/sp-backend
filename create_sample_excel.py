"""
Script to create a sample Excel file for product upload
Run this to generate sample_products.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_sample_excel():
    """Create a sample Excel file for product upload"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Headers
    headers = ['code', 'name', 'category', 'description', 'costPrice', 'sellingPrice', 'stockLevel', 'reorderLevel']
    ws.append(headers)
    
    # Format header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Sample data
    sample_products = [
        ['SKU001', 'Laptop', 'Electronics', 'High performance laptop', 45000, 75000, 10, 3],
        ['SKU002', 'Mouse', 'Accessories', 'Wireless mouse', 800, 1500, 50, 10],
        ['SKU003', 'Keyboard', 'Accessories', 'Mechanical keyboard', 2000, 3500, 30, 8],
        ['SKU004', 'Monitor', 'Electronics', '24 inch monitor', 8000, 14000, 5, 2],
        ['SKU005', 'USB Cable', 'Accessories', 'USB-C cable 2m', 200, 500, 100, 20],
        ['SKU006', 'Phone', 'Electronics', 'Smartphone', 20000, 35000, 15, 5],
        ['SKU007', 'Charger', 'Accessories', 'Fast charger', 1500, 2500, 40, 10],
        ['SKU008', 'Speaker', 'Electronics', 'Bluetooth speaker', 3000, 5500, 12, 3],
    ]
    
    # Add sample data with formatting
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_idx, row_data in enumerate(sample_products, start=2):
        ws.append(row_data)
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left" if col_idx <= 4 else "center")
    
    # Adjust column widths
    column_widths = [12, 20, 15, 25, 12, 14, 12, 14]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save file
    wb.save('sample_products.xlsx')
    print("✓ Sample Excel file created: sample_products.xlsx")
    print("\nYou can now:")
    print("1. Edit the sample data in the Excel file")
    print("2. Add more rows with your products")
    print("3. Upload to the API using the /api/products/upload endpoint")

if __name__ == '__main__':
    create_sample_excel()
