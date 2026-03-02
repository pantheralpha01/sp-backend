from openpyxl import load_workbook
from models import Product, db
import os

def parse_excel_products(file_path):
    """
    Parse Excel file and return list of product dictionaries
    
    Expected Excel columns: code, name, category, costPrice, sellingPrice, stockLevel, reorderLevel
    """
    products = []
    errors = []
    
    try:
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        
        # Get headers (first row)
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(cell.value.lower().strip())
        
        # Required columns
        required_cols = {'code', 'name', 'sellingprice'}
        found_cols = set(headers)
        
        if not required_cols.issubset(found_cols):
            missing = required_cols - found_cols
            return products, [f"Missing required columns: {', '.join(missing)}"]
        
        # Map headers to column indices
        col_map = {header: idx for idx, header in enumerate(headers)}
        
        # Process data rows (skip header)
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Skip empty rows
                if not row[col_map['code']] or not row[col_map['name']]:
                    continue
                
                product_data = {
                    'code': str(row[col_map['code']]).strip(),
                    'name': str(row[col_map['name']]).strip(),
                    'sellingPrice': float(row[col_map['sellingprice']] or 0),
                    'category': str(row[col_map.get('category', -1)] or 'General').strip(),
                    'costPrice': float(row[col_map.get('costprice', -1)] or 0),
                    'stockLevel': int(row[col_map.get('stocklevel', -1)] or 0),
                    'reorderLevel': int(row[col_map.get('reorderlevel', -1)] or 10),
                    'description': str(row[col_map.get('description', -1)] or '').strip(),
                }
                
                products.append(product_data)
                
            except (ValueError, TypeError) as e:
                errors.append(f"Row {row_idx}: Invalid data - {str(e)}")
                continue
        
        if not products:
            errors.append("No valid products found in Excel file")
        
        return products, errors
        
    except Exception as e:
        return products, [f"Error reading Excel file: {str(e)}"]

def save_products_to_db(products_data):
    """
    Save parsed products to database
    Returns: (success_count, duplicate_count, error_list)
    """
    success_count = 0
    duplicate_count = 0
    errors = []
    
    for product_data in products_data:
        try:
            # Check if product already exists
            existing = Product.query.filter_by(code=product_data['code']).first()
            
            if existing:
                duplicate_count += 1
                # Update existing product
                existing.name = product_data['name']
                existing.category = product_data['category']
                existing.description = product_data['description']
                existing.costPrice = product_data['costPrice']
                existing.sellingPrice = product_data['sellingPrice']
                existing.stockLevel = product_data['stockLevel']
                existing.reorderLevel = product_data['reorderLevel']
                db.session.add(existing)
            else:
                # Create new product
                product = Product(**product_data)
                db.session.add(product)
            
            success_count += 1
            
        except Exception as e:
            errors.append(f"Failed to save product {product_data.get('code', 'unknown')}: {str(e)}")
            continue
    
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return 0, 0, [f"Database error: {str(e)}"]
    
    return success_count, duplicate_count, errors
